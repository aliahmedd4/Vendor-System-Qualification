"""
Generates a minimal, valid OpenClinica 3.x CRF definition workbook:
  crf/VitalSigns-Min-CRF.xlsx

OpenClinica builds CRFs from an Excel workbook with four worksheets named exactly
CRF / Sections / Groups / Items, each with a fixed header row. This CRF defines a single
integer item (Systolic BP) with a hard range check 60-250 (VQ-004 CFG-046 / CFG-050), which
is enough to exercise: enter a value (OQ-11 create audit), edit it with a reason (OQ-12/14),
tamper-test the audit trail (OQ-15), and later the range edit check (OQ-20/21).

Run:  python scripts/build-crf.py
"""
from pathlib import Path
from openpyxl import Workbook

out = Path(__file__).resolve().parent.parent / "crf" / "VitalSigns-Min-CRF.xlsx"
out.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()

ws = wb.active
ws.title = "CRF"
ws.append(["CRF_NAME", "VERSION", "VERSION_DESCRIPTION", "REVISION_NOTES"])
ws.append(["Vital Signs (Min)", "v1.0",
           "Minimal single-field CRF for OQ qualification (VQ-004 CFG-046)", "initial"])

ws = wb.create_sheet("Sections")
ws.append(["SECTION_LABEL", "SECTION_TITLE", "SUBTITLE", "INSTRUCTIONS",
           "PAGE_NUMBER", "PARENT_SECTION"])
ws.append(["VITALS", "Vital Signs", "", "Enter the vital sign below.", "", ""])

ws = wb.create_sheet("Groups")
ws.append(["GROUP_LABEL", "GROUP_LAYOUT", "GROUP_HEADER",
           "GROUP_REPEAT_NUMBER", "GROUP_REPEAT_MAX", "GROUP_DISPLAY_STATUS"])
# (no repeating groups needed for a single item)

ws = wb.create_sheet("Items")
ws.append([
    "ITEM_NAME", "DESCRIPTION_LABEL", "LEFT_ITEM_TEXT", "UNITS", "RIGHT_ITEM_TEXT",
    "SECTION_LABEL", "GROUP_LABEL", "HEADER", "SUBHEADER", "PARENT_ITEM",
    "COLUMN_NUMBER", "PAGE_NUMBER", "QUESTION_NUMBER", "RESPONSE_TYPE", "RESPONSE_LABEL",
    "RESPONSE_OPTIONS_TEXT", "RESPONSE_VALUES_OR_CALCULATIONS", "RESPONSE_LAYOUT",
    "DEFAULT_VALUE", "DATA_TYPE", "WIDTH_DECIMAL", "VALIDATION", "VALIDATION_ERROR_MESSAGE",
    "PHI", "REQUIRED", "ITEM_DISPLAY_STATUS", "SIMPLE_CONDITIONAL_DISPLAY",
])
ws.append([
    "SYSBP", "Systolic Blood Pressure", "Systolic BP", "mmHg", "",
    "VITALS", "", "", "", "",
    1, "", "1", "text", "sysbp",
    "", "", "",
    "", "INT", "3(0)", "func: range(60, 250)", "Systolic BP must be between 60 and 250 mmHg",
    0, 1, "", "",
])

wb.save(out)
print("Wrote", out)
